import abc
import enum

from typing import Optional

from dataclasses import astuple
from datetime import datetime
from io import BytesIO

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Alignment, Border, Font, Side

from src.core.db.DTO_models import TasksAnalyticReportDto
from src.excel_generator.task_builder import AnalyticTaskReportFull


class AnalyticReportBuilder(abc.ABC):
    """Интерфейс строителя."""     
    def __add_row(self, analytic_task_report: Optional[AnalyticTaskReportFull],
                  data: tuple[str | int], worksheet: Optional[Worksheet]) -> None:
        analytic_task_report.row_count += 1
        for index, value in enumerate(data, start=1):
            worksheet.cell(row=analytic_task_report.row_count, column=index, value=value)

    async def __save_report(self, workbook: Workbook) -> BytesIO:
        """Сохранение отчёта."""
        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        return stream

    async def get_report_response(self, workbook: Workbook) -> StreamingResponse:
        """Создание ответа."""
        stream = await self.__save_report(workbook)
        filename = f"report_{datetime.now()}.xlsx"
        headers = {'Content-Disposition': f'attachment; filename={filename}'}
        return StreamingResponse(stream, headers=headers)

    def create_workbook(self) -> Workbook:
        """Генерация excel файла."""
        workbook = Workbook()
        workbook.remove_sheet(workbook.active)
        return workbook

    def create_sheet(self, workbook: Workbook, sheet_name: str) -> Workbook:
        """Создаёт лист внутри отчёта."""
        return workbook.create_sheet(sheet_name)

    def add_header(self, analytic_task_report: AnalyticTaskReportFull, worksheet: Worksheet) -> None:
        """Заполняет первые строки в листе."""
        for data in analytic_task_report.header_data:
            self.__add_row(
                analytic_task_report=analytic_task_report,
                data=data,
                worksheet=worksheet)

    def add_data(self, data: tuple[TasksAnalyticReportDto],
                 analytic_task_report: AnalyticTaskReportFull,
                 worksheet: Worksheet) -> None:
        """Заполняет строки данными из БД."""
        for task in data:
            self.__add_row(
                data=astuple(task),
                analytic_task_report=analytic_task_report,
                worksheet=worksheet)

    def add_footer(self, analytic_task_report: AnalyticTaskReportFull, worksheet: Worksheet) -> None:
        """Заполняет последнюю строку в листе."""
        self.__add_row(analytic_task_report=analytic_task_report,
                       data=analytic_task_report.footer_data,
                       worksheet=worksheet)

    def apply_styles(self, worksheet: Worksheet):
        """Задаёт форматирование отчёта."""
        # задаём стиль для хэдера
        rows = list(worksheet.iter_rows())
        header_cells = rows[0]
        for cell in header_cells:
            cell.font = self.Styles.FONT_BOLD.value
            cell.alignment = self.Styles.ALIGNMENT_HEADER.value
        # задаём стиль для ячеек с данными
        data_rows = rows[1:-1]
        for row in data_rows:
            for cell in row:
                cell.font = self.Styles.FONT_STANDART.value
                cell.alignment = self.Styles.ALIGNMENT_STANDART.value
        # задаём стиль для футера
        footer_cells = rows[-1]
        for cell in footer_cells:
            cell.font = self.Styles.FONT_BOLD.value
            cell.alignment = self.Styles.ALIGNMENT_STANDART.value
        # задаём стиль границ и колонок
        for row in rows:
            for cell in row:
                cell.border = self.Styles.BORDER.value
        worksheet.column_dimensions["A"].width = self.Styles.WIDTH.value

    class Styles(enum.Enum):
        FONT_BOLD = Font(name='Times New Roman', size=11, bold=True)
        FONT_STANDART = Font(name='Times New Roman', size=11, bold=False)
        ALIGNMENT_HEADER = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ALIGNMENT_STANDART = Alignment(horizontal='left', vertical='center', wrap_text=True)
        BORDER = Border(
            left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')
        )
        WIDTH = 50
